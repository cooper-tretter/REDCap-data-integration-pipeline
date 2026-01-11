"""
REDCap Data Integration Script
==============================

ONE ROW PER PARTICIPANT with comprehensive analytical tabs.

Features:
- Wide format: one row per participant, timepoint-specific columns
- Proper consent flow logic with status tracking
- _r timepoints consolidated with standard timepoints (t2_r -> t2, etc.)
- dosing_rescheduled flag preserved at participant level
- All major scales: PHQ-9, GAD-7, WHO-5, AUDIT-C, MEQ-4, PsyFlex, Expectancy
- Additional scales: EBI, PIQ, CEQ
- Multiple analytical tabs with outcomes and visualizations
- Treatment response analysis
- Mystical experience correlations

IMPORTANT: _r timepoint consolidation
- _r timepoints indicate the dosing session was rescheduled
- Data from _r timepoints is consolidated into the standard timepoint columns
  (e.g., phq9_total_t3 will contain data from timepoint_3_arm_1 OR timepoint_3_r_arm_1)
- The dosing_rescheduled flag indicates which participants had rescheduled sessions
- No participant should have both _r and non-_r data at the same logical timepoint
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
import sys
warnings.filterwarnings('ignore')


# =============================================================================
# TIMEPOINT MAPPING - Consolidates _r variants into standard timepoints
# =============================================================================

# Map both standard and _r events to the same logical timepoint
TIMEPOINT_MAP = {
    'timepoint_1_arm_1': 't1',
    'timepoint_2_arm_1': 't2',
    'timepoint_2_r_arm_1': 't2',    # _r consolidated to t2
    'timepoint_3_arm_1': 't3',
    'timepoint_3_r_arm_1': 't3',    # _r consolidated to t3
    'timepoint_4_arm_1': 't4',
    'timepoint_4_r_arm_1': 't4',    # _r consolidated to t4
    'timepoint_5_arm_1': 't5',
    'timepoint_5_r_arm_1': 't5',    # _r consolidated to t5
    'timepoint_6_arm_1': 't6',
    'timepoint_6_r_arm_1': 't6',    # _r consolidated to t6
}

TIMEPOINT_ORDER = ['t1', 't2', 't3', 't4', 't5', 't6']


# =============================================================================
# SCALE DEFINITIONS
# =============================================================================

SCALE_DEFINITIONS = {
    'phq9': {
        'name': 'PHQ-9 (Depression)',
        'items': 9,
        'item_range': (0, 3),
        'total_range': (0, 27),
        'scoring': 'sum',
        'interpretation': '0-4: None-minimal, 5-9: Mild, 10-14: Moderate, 15-19: Moderately severe, 20-27: Severe',
        'higher_is_worse': True,
        'events': 't1, t3, t4, t5, t6',  # NOT at t2 per protocol
    },
    'gad7': {
        'name': 'GAD-7 (Anxiety)',
        'items': 7,
        'item_range': (0, 3),
        'total_range': (0, 21),
        'scoring': 'sum',
        'interpretation': '0-4: Minimal, 5-9: Mild, 10-14: Moderate, 15-21: Severe',
        'higher_is_worse': True,
        'events': 't1, t3, t4, t5, t6',  # NOT at t2 per protocol
    },
    'who_5': {
        'name': 'WHO-5 (Wellbeing)',
        'items': 5,
        'item_range': (0, 5),
        'total_range': (0, 100),
        'scoring': 'sum_x4',
        'interpretation': '<28: Poor wellbeing, <13: Suggestive of depression, Higher = better',
        'higher_is_worse': False,
        'events': 't1, t3, t4, t5, t6',  # NOT at t2 per protocol
    },
    'psyflex': {
        'name': 'PsyFlex (Psychological Flexibility)',
        'items': 6,
        'item_range': (1, 5),
        'total_range': (6, 30),
        'scoring': 'sum',
        'interpretation': 'Higher scores = greater psychological flexibility',
        'higher_is_worse': False,
        'events': 't1, t3, t4, t5, t6',
    },
    'auditc': {
        'name': 'AUDIT-C (Alcohol Use)',
        'items': 3,
        'item_range': (0, 4),
        'total_range': (0, 12),
        'scoring': 'sum',
        'interpretation': '>=4 (men) or >=3 (women): Hazardous drinking',
        'higher_is_worse': True,
        'events': 't1, t3, t4, t5, t6',
    },
    'meq4': {
        'name': 'MEQ-4 (Mystical Experience)',
        'items': 4,
        'item_range': (0, 5),
        'total_range': (0, 5),
        'scoring': 'mean',
        'interpretation': '>=3.5: Complete mystical experience',
        'higher_is_worse': False,
        'events': 't2 only',  # Only at dosing session
    },
    'expect': {
        'name': 'Expectancy',
        'items': 1,
        'item_range': (0, 10),
        'total_range': (0, 10),
        'scoring': 'single',
        'interpretation': 'Higher = greater treatment expectancy',
        'higher_is_worse': False,
        'events': 't1 only',
    },
    'ebi': {
        'name': 'EBI (Emotional Breakthrough)',
        'items': 6,
        'item_range': (0, 5),
        'total_range': (0, 30),
        'scoring': 'sum',
        'interpretation': 'Higher = greater emotional breakthrough during session',
        'higher_is_worse': False,
        'events': 't2 only',
    },
    'piq': {
        'name': 'PIQ (Psychological Insight)',
        'items': 9,
        'item_range': (1, 5),
        'total_range': (9, 45),
        'scoring': 'sum',
        'interpretation': 'Higher = greater psychological insight',
        'higher_is_worse': False,
        'events': 't2 only',
    },
    'ceq': {
        'name': 'CEQ (Challenging Experience)',
        'items': 7,
        'item_range': (0, 5),
        'total_range': (0, 35),
        'scoring': 'sum',
        'interpretation': 'Higher = more challenging/difficult experience',
        'higher_is_worse': True,
        'events': 't2 only',
    },
}


# =============================================================================
# SCORING FUNCTIONS
# =============================================================================

def calc_sum(row, items):
    """Sum of items, all must be present."""
    vals = [row[i] for i in items if i in row.index and pd.notna(row[i])]
    return sum(vals) if len(vals) == len(items) else np.nan


def calc_mean(row, items):
    """Mean of items, all must be present."""
    vals = [row[i] for i in items if i in row.index and pd.notna(row[i])]
    return np.mean(vals) if len(vals) == len(items) else np.nan


def phq9_severity(score):
    if pd.isna(score): return np.nan
    elif score < 5: return 'None-minimal'
    elif score < 10: return 'Mild'
    elif score < 15: return 'Moderate'
    elif score < 20: return 'Moderately severe'
    else: return 'Severe'


def gad7_severity(score):
    if pd.isna(score): return np.nan
    elif score < 5: return 'Minimal'
    elif score < 10: return 'Mild'
    elif score < 15: return 'Moderate'
    else: return 'Severe'


def collapse_checkbox(df, prefix, n_options, labels=None):
    """Collapse checkbox fields to single column."""
    if labels is None:
        labels = {i: f'Option {i}' for i in range(1, n_options + 1)}

    def collapse_row(row):
        selected = [labels[i] for i in range(1, n_options + 1)
                   if f'{prefix}{i}' in row.index and row[f'{prefix}{i}'] == 1]
        return ', '.join(selected) if selected else np.nan

    return df.apply(collapse_row, axis=1)


# =============================================================================
# DETERMINE IF PARTICIPANT HAD RESCHEDULED DOSING
# =============================================================================

def determine_dosing_rescheduled(df):
    """
    Determine if each participant had a rescheduled dosing session.

    A participant is considered to have a rescheduled session if they have
    any _r timepoint in their data.

    Returns:
        DataFrame with record_id and dosing_rescheduled columns
    """
    rescheduled_info = []

    for record_id in df['record_id'].unique():
        participant_events = df[df['record_id'] == record_id]['redcap_event_name'].tolist()
        has_r_events = any('_r_' in event for event in participant_events)
        rescheduled_info.append({
            'record_id': record_id,
            'dosing_rescheduled': has_r_events
        })

    return pd.DataFrame(rescheduled_info)


# =============================================================================
# EXTRACT PARTICIPANT INFO WITH CONSENT LOGIC
# =============================================================================

def extract_participant_info(df, df_rescheduled):
    """
    Extract participant info respecting consent branching logic.
    """
    print("   Extracting participant info with consent logic...")

    all_record_ids = df['record_id'].unique()
    participant_info = []

    for record_id in all_record_ids:
        participant_rows = df[df['record_id'] == record_id]
        info = {'record_id': record_id}

        # Get rescheduled status
        rescheduled_row = df_rescheduled[df_rescheduled['record_id'] == record_id]
        info['dosing_rescheduled'] = rescheduled_row['dosing_rescheduled'].iloc[0] if len(rescheduled_row) > 0 else False

        # Check if participant has baseline
        baseline_rows = participant_rows[participant_rows['redcap_event_name'] == 'timepoint_1_arm_1']
        has_baseline = len(baseline_rows) > 0
        info['has_baseline'] = has_baseline

        if has_baseline:
            baseline = baseline_rows.iloc[0]

            consent_age = baseline.get('consent_age', np.nan)
            info['consent_age'] = consent_age

            if pd.isna(consent_age):
                info['consent_status'] = 'incomplete'
                info['consent_passed'] = False
            elif consent_age == 0:
                info['consent_status'] = 'failed_age_check'
                info['consent_passed'] = False
            else:
                psilocybin = baseline.get('consent_psilocybintherapy', np.nan)
                info['consent_psilocybintherapy'] = psilocybin

                if pd.isna(psilocybin):
                    info['consent_status'] = 'incomplete'
                    info['consent_passed'] = False
                elif psilocybin == 0:
                    info['consent_status'] = 'failed_psilocybin_check'
                    info['consent_passed'] = False
                else:
                    # Check if any consent version has a name
                    has_any_name = False
                    for name_field in ['consent_nameprint', 'consent_nameprint_v2', 'consent_nameprint_v3']:
                        if name_field in baseline.index:
                            val = baseline[name_field]
                            if pd.notna(val) and str(val).strip() != '':
                                has_any_name = True
                                break

                    if has_any_name:
                        info['consent_status'] = 'passed'
                        info['consent_passed'] = True
                    else:
                        info['consent_status'] = 'eligible_but_incomplete'
                        info['consent_passed'] = False

            # Extract name from v1, v2, or v3
            name = None
            name_source = None
            for name_field, source in [('consent_nameprint', 'v1'),
                                       ('consent_nameprint_v2', 'v2'),
                                       ('consent_nameprint_v3', 'v3')]:
                if name_field in baseline.index:
                    val = baseline[name_field]
                    if pd.notna(val) and str(val).strip() != '':
                        name = val
                        name_source = source
                        break

            info['consent_nameprint'] = name
            info['name_source'] = name_source

            # Extract email similarly
            email = None
            for email_field in ['email', 'email_v2', 'email_v3']:
                if email_field in baseline.index and pd.notna(baseline[email_field]):
                    email = baseline[email_field]
                    break
            info['email'] = email

            # Extract demographics
            demo_vars = ['age', 'gender', 'sex', 'education', 'relat', 'latino',
                        'income_est', 'military_service', 'consent_date']
            for var in demo_vars:
                if var in baseline.index:
                    info[var] = baseline[var]

            # Extract checkbox fields
            checkbox_prefixes = [
                ('race1___', 6),
                ('employ___', 9),
                ('psychiatric_medications___', 8),
                ('psychedelics_used___', 9),
            ]
            for prefix, n in checkbox_prefixes:
                for i in range(1, n + 1):
                    col = f'{prefix}{i}'
                    if col in baseline.index:
                        info[col] = baseline[col]
        else:
            info['consent_age'] = np.nan
            info['consent_status'] = 'no_baseline'
            info['consent_passed'] = False
            info['consent_nameprint'] = np.nan
            info['name_source'] = None
            info['email'] = np.nan

        # Track which events the participant has (original event names)
        info['events_original'] = ', '.join(participant_rows['redcap_event_name'].tolist())
        info['n_events'] = len(participant_rows)

        # Track consolidated timepoints
        consolidated_tps = sorted(set([TIMEPOINT_MAP.get(e, e) for e in participant_rows['redcap_event_name'].tolist()]))
        info['timepoints'] = ', '.join(consolidated_tps)

        participant_info.append(info)

    df_participants = pd.DataFrame(participant_info)

    print(f"   - Total participants: {len(df_participants)}")
    print(f"   - With baseline: {df_participants['has_baseline'].sum()}")
    print(f"   - Consent passed: {df_participants['consent_passed'].sum()}")
    print(f"   - With name: {df_participants['consent_nameprint'].notna().sum()}")
    print(f"   - Dosing rescheduled: {df_participants['dosing_rescheduled'].sum()}")

    return df_participants


# =============================================================================
# PIVOT TIME-VARYING DATA (WITH _R CONSOLIDATION)
# =============================================================================

def pivot_time_varying(df):
    """
    Pivot time-varying variables to wide format.

    IMPORTANT: _r timepoints are consolidated into standard timepoints.
    Data from timepoint_3_r_arm_1 goes into t3 columns, not t3_r columns.
    """
    print("   Pivoting time-varying variables (consolidating _r timepoints)...")

    # Build list of time-varying columns
    time_varying = []

    # Core scales
    for scale in ['phq9', 'gad7']:
        time_varying.extend([f'{scale}_{i}' for i in range(1, 10)])
        time_varying.append(f'{scale}_totalscore')

    time_varying.extend([f'who_5_{i}' for i in range(1, 6)])
    time_varying.extend([f'psyflex_{i}' for i in range(1, 7)])
    time_varying.extend([f'auditc_{i}' for i in range(1, 4)])
    time_varying.extend([f'meq4_{i}' for i in range(1, 5)])

    # Additional scales
    time_varying.extend([f'ebi_{i}' for i in range(1, 7)])
    time_varying.extend([f'piq_{i}' for i in range(1, 10)])
    time_varying.extend([f'ceq_{i}' for i in range(1, 8)])

    # Other fields
    time_varying.extend(['expect_1', 'treatment_date', 'treatment_status'])

    # Filter to columns that exist
    time_varying = [c for c in time_varying if c in df.columns]

    # Add consolidated timepoint column
    df_wide = df[['record_id', 'redcap_event_name'] + time_varying].copy()
    df_wide['timepoint'] = df_wide['redcap_event_name'].map(TIMEPOINT_MAP)

    # Pivot - since _r is consolidated, there should be no duplicates
    pivoted_dfs = []
    for var in time_varying:
        try:
            df_var = df_wide[['record_id', 'timepoint', var]].pivot(
                index='record_id', columns='timepoint', values=var
            )
            df_var.columns = [f'{var}_{col}' for col in df_var.columns]
            pivoted_dfs.append(df_var.reset_index())
        except Exception as e:
            # If pivot fails (e.g., duplicates), use groupby with first value
            df_grouped = df_wide.groupby(['record_id', 'timepoint'])[var].first().unstack()
            df_grouped.columns = [f'{var}_{col}' for col in df_grouped.columns]
            pivoted_dfs.append(df_grouped.reset_index())

    if pivoted_dfs:
        df_merged = pivoted_dfs[0]
        for pivoted in pivoted_dfs[1:]:
            df_merged = df_merged.merge(pivoted, on='record_id', how='outer')
        return df_merged
    else:
        return pd.DataFrame({'record_id': df['record_id'].unique()})


# =============================================================================
# CALCULATE DERIVED SCORES
# =============================================================================

def calculate_all_scores(df):
    """Calculate all scale scores at each timepoint."""
    print("   Computing scores for all timepoints...")

    for tp in TIMEPOINT_ORDER:
        # PHQ-9
        phq9_items = [f'phq9_{i}_{tp}' for i in range(1, 10)]
        if all(c in df.columns for c in phq9_items):
            df[f'phq9_total_{tp}'] = df.apply(lambda row: calc_sum(row, phq9_items), axis=1)
            df[f'phq9_severity_{tp}'] = df[f'phq9_total_{tp}'].apply(phq9_severity)
        elif f'phq9_totalscore_{tp}' in df.columns:
            df[f'phq9_total_{tp}'] = df[f'phq9_totalscore_{tp}']
            df[f'phq9_severity_{tp}'] = df[f'phq9_total_{tp}'].apply(phq9_severity)

        # GAD-7
        gad7_items = [f'gad7_{i}_{tp}' for i in range(1, 8)]
        if all(c in df.columns for c in gad7_items):
            df[f'gad7_total_{tp}'] = df.apply(lambda row: calc_sum(row, gad7_items), axis=1)
            df[f'gad7_severity_{tp}'] = df[f'gad7_total_{tp}'].apply(gad7_severity)
        elif f'gad7_totalscore_{tp}' in df.columns:
            df[f'gad7_total_{tp}'] = df[f'gad7_totalscore_{tp}']
            df[f'gad7_severity_{tp}'] = df[f'gad7_total_{tp}'].apply(gad7_severity)

        # WHO-5
        who5_items = [f'who_5_{i}_{tp}' for i in range(1, 6)]
        if all(c in df.columns for c in who5_items):
            raw_sum = df.apply(lambda row: calc_sum(row, who5_items), axis=1)
            df[f'who5_total_{tp}'] = raw_sum * 4

        # PsyFlex
        psyflex_items = [f'psyflex_{i}_{tp}' for i in range(1, 7)]
        if all(c in df.columns for c in psyflex_items):
            df[f'psyflex_total_{tp}'] = df.apply(lambda row: calc_sum(row, psyflex_items), axis=1)

        # AUDIT-C
        auditc_items = [f'auditc_{i}_{tp}' for i in range(1, 4)]
        if all(c in df.columns for c in auditc_items):
            df[f'auditc_total_{tp}'] = df.apply(lambda row: calc_sum(row, auditc_items), axis=1)

        # MEQ-4
        meq4_items = [f'meq4_{i}_{tp}' for i in range(1, 5)]
        if all(c in df.columns for c in meq4_items):
            df[f'meq4_mean_{tp}'] = df.apply(lambda row: calc_mean(row, meq4_items), axis=1)
            df[f'meq4_mystical_{tp}'] = df[f'meq4_mean_{tp}'].apply(
                lambda x: 'Yes' if pd.notna(x) and x >= 3.5 else ('No' if pd.notna(x) else np.nan)
            )

        # EBI (Emotional Breakthrough)
        ebi_items = [f'ebi_{i}_{tp}' for i in range(1, 7)]
        available_ebi = [c for c in ebi_items if c in df.columns]
        if len(available_ebi) >= 4:
            df[f'ebi_total_{tp}'] = df.apply(
                lambda row: calc_sum(row, available_ebi) if sum(pd.notna(row[c]) for c in available_ebi) == len(available_ebi) else np.nan,
                axis=1
            )

        # PIQ (Psychological Insight)
        piq_items = [f'piq_{i}_{tp}' for i in range(1, 10)]
        if all(c in df.columns for c in piq_items):
            df[f'piq_total_{tp}'] = df.apply(lambda row: calc_sum(row, piq_items), axis=1)

        # CEQ (Challenging Experience)
        ceq_items = [f'ceq_{i}_{tp}' for i in range(1, 8)]
        if all(c in df.columns for c in ceq_items):
            df[f'ceq_total_{tp}'] = df.apply(lambda row: calc_sum(row, ceq_items), axis=1)

        # Expectancy (single item)
        if f'expect_1_{tp}' in df.columns:
            df[f'expect_total_{tp}'] = df[f'expect_1_{tp}']

    return df


# =============================================================================
# ANALYTICAL TAB FUNCTIONS
# =============================================================================

def create_scale_summary(df, scale_name):
    """Create summary statistics for a scale across timepoints."""
    summary_data = []

    for tp in TIMEPOINT_ORDER:
        col = f'{scale_name}_{tp}'
        if col in df.columns:
            n = df[col].notna().sum()
            if n > 0:
                summary_data.append({
                    'Timepoint': tp,
                    'N': n,
                    'Mean': round(df[col].mean(), 2),
                    'SD': round(df[col].std(), 2),
                    'Median': round(df[col].median(), 2),
                    'Min': df[col].min(),
                    'Max': df[col].max(),
                })

    return pd.DataFrame(summary_data)


def create_improvement_analysis(df, scale_name, baseline='t1', higher_is_worse=True):
    """Analyze improvement from baseline to follow-ups."""
    improvement_data = []

    baseline_col = f'{scale_name}_{baseline}'
    if baseline_col not in df.columns:
        return pd.DataFrame()

    for tp in ['t3', 't4', 't5', 't6']:  # Note: t2 excluded for PHQ-9/GAD-7/WHO-5
        tp_col = f'{scale_name}_{tp}'
        if tp_col in df.columns:
            both = df[baseline_col].notna() & df[tp_col].notna()
            if both.sum() > 0:
                df_paired = df[both].copy()
                df_paired['change'] = df_paired[tp_col] - df_paired[baseline_col]

                if higher_is_worse:
                    df_paired['responder'] = df_paired['change'] <= -0.5 * df_paired[baseline_col]
                    df_paired['improved'] = df_paired['change'] < 0
                else:
                    df_paired['responder'] = df_paired['change'] >= 0.5 * df_paired[baseline_col]
                    df_paired['improved'] = df_paired['change'] > 0

                improvement_data.append({
                    'Comparison': f'{baseline} to {tp}',
                    'N_paired': len(df_paired),
                    'Mean_baseline': round(df_paired[baseline_col].mean(), 2),
                    'Mean_followup': round(df_paired[tp_col].mean(), 2),
                    'Mean_change': round(df_paired['change'].mean(), 2),
                    'SD_change': round(df_paired['change'].std(), 2),
                    'Improved_N': df_paired['improved'].sum(),
                    'Improved_pct': round(df_paired['improved'].mean() * 100, 1),
                    'Responders_N': df_paired['responder'].sum(),
                    'Responders_pct': round(df_paired['responder'].mean() * 100, 1),
                })

    return pd.DataFrame(improvement_data)


def create_meq_analysis(df):
    """Analyze mystical experiences and correlations."""
    print("   Analyzing mystical experiences...")

    summary_data = []
    mean_col = 'meq4_mean_t2'
    myst_col = 'meq4_mystical_t2'

    if mean_col in df.columns:
        n = df[mean_col].notna().sum()
        if n > 0:
            n_mystical = (df[myst_col] == 'Yes').sum() if myst_col in df.columns else 0
            pct_mystical = (n_mystical / n * 100) if n > 0 else 0

            summary_data.append({
                'Timepoint': 't2 (dosing)',
                'N_completed': n,
                'Mean_score': round(df[mean_col].mean(), 2),
                'SD': round(df[mean_col].std(), 2),
                'N_mystical_exp': n_mystical,
                'Pct_mystical_exp': round(pct_mystical, 1),
            })

    df_meq_summary = pd.DataFrame(summary_data)

    # Correlations with outcomes
    correlation_data = []
    meq_col = 'meq4_mean_t2'

    for outcome in ['phq9_total', 'gad7_total', 'who5_total']:
        baseline_col = f'{outcome}_t1'
        followup_col = f'{outcome}_t3'

        if meq_col in df.columns and baseline_col in df.columns and followup_col in df.columns:
            df_temp = df.copy()
            df_temp['change'] = df_temp[followup_col] - df_temp[baseline_col]

            valid = df_temp[meq_col].notna() & df_temp['change'].notna()
            if valid.sum() >= 3:
                corr = df_temp.loc[valid, meq_col].corr(df_temp.loc[valid, 'change'])
                correlation_data.append({
                    'MEQ_Timepoint': 't2',
                    'Outcome': outcome,
                    'Outcome_Change': 't1 to t3',
                    'Correlation_r': round(corr, 3) if pd.notna(corr) else np.nan,
                    'N': valid.sum(),
                })

    df_correlations = pd.DataFrame(correlation_data)
    return df_meq_summary, df_correlations


def create_acute_experience_analysis(df):
    """Analyze acute experience measures (EBI, PIQ, CEQ)."""
    print("   Analyzing acute experience measures...")

    summary_data = []

    for scale, scale_name in [('ebi_total', 'EBI (Emotional Breakthrough)'),
                               ('piq_total', 'PIQ (Psychological Insight)'),
                               ('ceq_total', 'CEQ (Challenging Experience)')]:
        col = f'{scale}_t2'
        if col in df.columns:
            n = df[col].notna().sum()
            if n > 0:
                summary_data.append({
                    'Scale': scale_name,
                    'Timepoint': 't2 (dosing)',
                    'N': n,
                    'Mean': round(df[col].mean(), 2),
                    'SD': round(df[col].std(), 2),
                    'Median': round(df[col].median(), 2),
                    'Min': df[col].min(),
                    'Max': df[col].max(),
                })

    return pd.DataFrame(summary_data)


def create_demographics_summary(df):
    """Create demographics summary for consented participants."""
    print("   Creating demographics summary...")

    df_consented = df[df['consent_passed'] == True].copy()

    if len(df_consented) == 0:
        return pd.DataFrame({'Note': ['No consented participants found']})

    summary = {'Variable': [], 'Category': [], 'N': [], 'Percent': []}

    # Age
    if 'age' in df_consented.columns:
        age_valid = df_consented['age'].dropna()
        if len(age_valid) > 0:
            summary['Variable'].append('Age')
            summary['Category'].append('Mean (SD)')
            summary['N'].append(f"{age_valid.mean():.1f} ({age_valid.std():.1f})")
            summary['Percent'].append('')

    # Dosing rescheduled
    if 'dosing_rescheduled' in df_consented.columns:
        rescheduled = df_consented['dosing_rescheduled'].sum()
        total = len(df_consented)
        summary['Variable'].append('Dosing Rescheduled')
        summary['Category'].append('Yes')
        summary['N'].append(rescheduled)
        summary['Percent'].append(f"{100*rescheduled/total:.1f}%" if total > 0 else '')

    # Gender
    if 'gender' in df_consented.columns:
        for code, label in {1: 'Male', 2: 'Female', 3: 'Non-binary', 4: 'Other'}.items():
            count = (df_consented['gender'] == code).sum()
            if count > 0:
                summary['Variable'].append('Gender')
                summary['Category'].append(label)
                summary['N'].append(count)
                total = df_consented['gender'].notna().sum()
                summary['Percent'].append(f"{100*count/total:.1f}%" if total > 0 else '')

    # Education
    if 'education' in df_consented.columns:
        edu_labels = {1: 'Less than HS', 2: 'High School', 3: 'Some College',
                     4: 'Bachelor\'s', 5: 'Graduate'}
        for code, label in edu_labels.items():
            count = (df_consented['education'] == code).sum()
            if count > 0:
                summary['Variable'].append('Education')
                summary['Category'].append(label)
                summary['N'].append(count)
                total = df_consented['education'].notna().sum()
                summary['Percent'].append(f"{100*count/total:.1f}%" if total > 0 else '')

    return pd.DataFrame(summary)


def create_completeness_summary(df):
    """Data completeness across timepoints."""
    print("   Creating completeness summary...")

    scales = ['phq9_total', 'gad7_total', 'who5_total', 'psyflex_total',
              'auditc_total', 'meq4_mean', 'ebi_total', 'piq_total', 'ceq_total']

    completeness_data = []
    for tp in TIMEPOINT_ORDER:
        row = {'Timepoint': tp}
        for scale in scales:
            col = f'{scale}_{tp}'
            if col in df.columns:
                row[scale] = df[col].notna().sum()
            else:
                row[scale] = 0
        completeness_data.append(row)

    return pd.DataFrame(completeness_data)


def create_summary_tab(df):
    """Create concise summary with key demographics and first/last survey totals."""
    print("   Creating summary tab...")

    summary_data = []

    scales = {
        'phq9_total': 'PHQ9',
        'gad7_total': 'GAD7',
        'who5_total': 'WHO5',
        'meq4_mean': 'MEQ4',
        'psyflex_total': 'PsyFlex',
        'auditc_total': 'AUDIT-C',
        'ebi_total': 'EBI',
        'piq_total': 'PIQ',
        'ceq_total': 'CEQ',
    }

    for _, row in df.iterrows():
        summary_row = {
            'record_id': row['record_id'],
            'consent_nameprint': row.get('consent_nameprint', np.nan),
            'consent_status': row.get('consent_status', np.nan),
            'consent_passed': row.get('consent_passed', np.nan),
            'dosing_rescheduled': row.get('dosing_rescheduled', np.nan),
            'age': row.get('age', np.nan),
            'gender': row.get('gender', np.nan),
            'n_events': row.get('n_events', np.nan),
            'timepoints': row.get('timepoints', np.nan),
        }

        for scale_base, scale_label in scales.items():
            first_val = np.nan
            first_tp = None
            last_val = np.nan
            last_tp = None

            for tp in TIMEPOINT_ORDER:
                col = f'{scale_base}_{tp}'
                if col in row.index and pd.notna(row[col]):
                    if pd.isna(first_val):
                        first_val = row[col]
                        first_tp = tp
                    last_val = row[col]
                    last_tp = tp

            summary_row[f'{scale_base}_first'] = first_val
            summary_row[f'{scale_base}_first_tp'] = first_tp
            summary_row[f'{scale_base}_last'] = last_val
            summary_row[f'{scale_base}_last_tp'] = last_tp

            if pd.notna(first_val) and pd.notna(last_val) and first_tp != last_tp:
                summary_row[f'{scale_base}_change'] = last_val - first_val
                if first_val != 0:
                    summary_row[f'{scale_base}_pct_change'] = ((last_val - first_val) / first_val) * 100
                else:
                    summary_row[f'{scale_base}_pct_change'] = np.nan
            else:
                summary_row[f'{scale_base}_change'] = np.nan
                summary_row[f'{scale_base}_pct_change'] = np.nan

        summary_data.append(summary_row)

    df_summary = pd.DataFrame(summary_data)

    first_cols = [f'{s}_first' for s in scales.keys()]
    last_cols = [f'{s}_last' for s in scales.keys()]
    change_cols = [f'{s}_change' for s in scales.keys()]

    return df_summary, first_cols, last_cols, change_cols


def create_calculations_tab():
    """Create documentation tab explaining how all scores are calculated."""

    calculations = []
    for scale_key, scale_info in SCALE_DEFINITIONS.items():
        calculations.append({
            'Measure': scale_info['name'],
            'Score Name': f'{scale_key}_total' if scale_info['scoring'] != 'mean' else f'{scale_key}_mean',
            'Calculation': f"{scale_info['scoring'].replace('_', ' ').title()} of {scale_info['items']} items",
            'Item Range': f"{scale_info['item_range'][0]}-{scale_info['item_range'][1]} each",
            'Total Range': f"{scale_info['total_range'][0]}-{scale_info['total_range'][1]}",
            'Events Available': scale_info['events'],
            'Interpretation': scale_info['interpretation'],
        })

    # Add response/remission definitions
    calculations.extend([
        {
            'Measure': 'Clinical Response',
            'Score Name': 'Responder status',
            'Calculation': '>=50% reduction from baseline',
            'Item Range': 'N/A',
            'Total Range': 'Yes/No',
            'Events Available': 'Calculated',
            'Interpretation': 'Standard clinical trial criterion for meaningful symptom improvement',
        },
        {
            'Measure': 'Remission',
            'Score Name': 'Remission status',
            'Calculation': 'Final score <5 (PHQ-9/GAD-7) or >=50 (WHO-5)',
            'Item Range': 'N/A',
            'Total Range': 'Yes/No',
            'Events Available': 'Calculated',
            'Interpretation': 'Below clinical cutoff indicating minimal symptoms',
        },
    ])

    df_calc = pd.DataFrame(calculations)

    notes = [
        {'Note Type': 'Data Structure', 'Description': 'One row per participant with timepoint-specific columns (e.g., phq9_total_t1)'},
        {'Note Type': 'Timepoint Codes', 'Description': 't1=Baseline, t2=Dosing session, t3-t6=Follow-ups'},
        {'Note Type': '_r Consolidation', 'Description': '_r timepoints (rescheduled dosing) are consolidated into standard timepoints. Check dosing_rescheduled column to identify rescheduled participants.'},
        {'Note Type': 'Questionnaire Timing', 'Description': 'PHQ-9/GAD-7/WHO-5/PsyFlex/AUDIT-C at t1,t3-t6 (NOT t2). MEQ/EBI/CEQ/PIQ at t2 only.'},
        {'Note Type': 'Consent Status', 'Description': 'passed=completed, failed_age_check=under 21, failed_psilocybin_check=not seeking therapy, incomplete=didn\'t finish'},
        {'Note Type': 'Missing Data', 'Description': 'All scoring requires complete item responses. Partial responses result in missing total scores.'},
    ]

    df_notes = pd.DataFrame(notes)

    return df_calc, df_notes


# =============================================================================
# MAIN INTEGRATION
# =============================================================================

def integrate_full(input_path, output_dir=None):
    """Full integration with comprehensive analytics."""

    print("=" * 80)
    print("REDCap Data Integration")
    print("ONE ROW PER PARTICIPANT - _r TIMEPOINTS CONSOLIDATED")
    print("=" * 80)

    # Read data
    print(f"\n1. Reading data: {input_path}")
    df = pd.read_excel(input_path) if str(input_path).endswith('.xlsx') else pd.read_csv(input_path)
    print(f"   - Original: {len(df)} rows x {len(df.columns)} columns")
    print(f"   - Unique participants: {df['record_id'].nunique()}")

    # Check for _r events
    r_events = df[df['redcap_event_name'].str.contains('_r_', na=False)]
    print(f"   - Rows with _r timepoints: {len(r_events)}")

    output_dir = Path(output_dir or Path(input_path).parent)

    # Step 0: Determine dosing rescheduled status
    print("\n2. Determining dosing rescheduled status...")
    df_rescheduled = determine_dosing_rescheduled(df)
    n_rescheduled = df_rescheduled['dosing_rescheduled'].sum()
    print(f"   - Participants with rescheduled dosing: {n_rescheduled}")

    # Step 1: Extract participant info with consent logic
    print("\n3. Extracting participant info...")
    df_participants = extract_participant_info(df, df_rescheduled)

    # Step 2: Pivot time-varying data (with _r consolidation)
    print("\n4. Pivoting time-varying data...")
    df_time_varying = pivot_time_varying(df)

    # Step 3: Merge
    print("\n5. Merging data...")
    df_wide = df_participants.merge(df_time_varying, on='record_id', how='left')
    print(f"   - Result: {len(df_wide)} rows x {len(df_wide.columns)} columns")

    # Step 4: Collapse checkboxes
    print("\n6. Collapsing checkbox fields...")
    checkbox_defs = {
        'race1___': (6, {1: 'AI/AN', 2: 'Asian', 3: 'Black', 4: 'NH/PI', 5: 'White', 6: 'Other'}),
        'employ___': (9, None),
        'psychiatric_medications___': (8, None),
        'psychedelics_used___': (9, {1: 'Psilocybin', 2: 'LSD', 3: 'MDMA', 4: 'Ayahuasca',
                                     5: 'DMT', 6: 'Mescaline', 7: 'Ketamine', 8: 'Salvia', 9: 'Other'}),
    }

    for prefix, (n, labels) in checkbox_defs.items():
        cols = [f'{prefix}{i}' for i in range(1, n + 1)]
        if any(c in df_wide.columns for c in cols):
            new_name = prefix.rstrip('_').replace('___', '')
            df_wide[new_name] = collapse_checkbox(df_wide, prefix, n, labels)
            df_wide.drop(columns=[c for c in cols if c in df_wide.columns], inplace=True, errors='ignore')

    # Step 5: Calculate scores
    print("\n7. Computing all scores...")
    df_wide = calculate_all_scores(df_wide)

    # Step 6: Create analytical tabs
    print("\n8. Creating analytical tabs...")

    df_summary, first_cols, last_cols, change_cols = create_summary_tab(df_wide)
    df_demographics = create_demographics_summary(df_wide)
    df_completeness = create_completeness_summary(df_wide)

    # Scale-specific tabs
    df_phq9_summary = create_scale_summary(df_wide, 'phq9_total')
    df_phq9_outcomes = create_improvement_analysis(df_wide, 'phq9_total', higher_is_worse=True)

    df_gad7_summary = create_scale_summary(df_wide, 'gad7_total')
    df_gad7_outcomes = create_improvement_analysis(df_wide, 'gad7_total', higher_is_worse=True)

    df_who5_summary = create_scale_summary(df_wide, 'who5_total')
    df_who5_outcomes = create_improvement_analysis(df_wide, 'who5_total', higher_is_worse=False)

    df_psyflex_summary = create_scale_summary(df_wide, 'psyflex_total')
    df_auditc_summary = create_scale_summary(df_wide, 'auditc_total')

    df_meq_summary, df_meq_correlations = create_meq_analysis(df_wide)
    df_acute = create_acute_experience_analysis(df_wide)

    df_calculations, df_calc_notes = create_calculations_tab()

    # Step 7: Save outputs
    print("\n9. Saving outputs...")
    timestamp = datetime.now().strftime("%Y%m%d_T%H%M%S")
    excel_output = output_dir / f'insights_{timestamp}.xlsx'

    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        df_wide.to_excel(writer, sheet_name='Main Data', index=False)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_demographics.to_excel(writer, sheet_name='Demographics', index=False)
        df_completeness.to_excel(writer, sheet_name='Data Completeness', index=False)

        if not df_phq9_summary.empty:
            df_phq9_summary.to_excel(writer, sheet_name='PHQ9 Summary', index=False)
        if not df_phq9_outcomes.empty:
            df_phq9_outcomes.to_excel(writer, sheet_name='PHQ9 Outcomes', index=False)

        if not df_gad7_summary.empty:
            df_gad7_summary.to_excel(writer, sheet_name='GAD7 Summary', index=False)
        if not df_gad7_outcomes.empty:
            df_gad7_outcomes.to_excel(writer, sheet_name='GAD7 Outcomes', index=False)

        if not df_who5_summary.empty:
            df_who5_summary.to_excel(writer, sheet_name='WHO5 Summary', index=False)
        if not df_who5_outcomes.empty:
            df_who5_outcomes.to_excel(writer, sheet_name='WHO5 Outcomes', index=False)

        if not df_psyflex_summary.empty:
            df_psyflex_summary.to_excel(writer, sheet_name='PsyFlex Summary', index=False)

        if not df_auditc_summary.empty:
            df_auditc_summary.to_excel(writer, sheet_name='AUDIT-C Summary', index=False)

        if not df_meq_summary.empty:
            df_meq_summary.to_excel(writer, sheet_name='MEQ Analysis', index=False)
        if not df_meq_correlations.empty:
            df_meq_correlations.to_excel(writer, sheet_name='MEQ Correlations', index=False)

        if not df_acute.empty:
            df_acute.to_excel(writer, sheet_name='Acute Experience', index=False)

        df_calculations.to_excel(writer, sheet_name='Calculations', index=False, startrow=0)
        df_calc_notes.to_excel(writer, sheet_name='Calculations', index=False,
                               startrow=len(df_calculations) + 3)

    # Apply color formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(excel_output)
    ws = wb['Summary']

    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    header_row = [cell.value for cell in ws[1]]

    for col_name in first_cols:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            ws.cell(row=1, column=col_idx).fill = orange_fill

    for col_name in last_cols:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            ws.cell(row=1, column=col_idx).fill = green_fill

    wb.save(excel_output)
    print(f"   - Saved: {excel_output}")

    csv_output = output_dir / f'insights_{timestamp}.csv'
    df_wide.to_csv(csv_output, index=False)
    print(f"   - Saved: {csv_output}")

    # Final report
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Input: {len(df)} rows (long format)")
    print(f"Output: {len(df_wide)} rows (wide format - one per participant)")
    print(f"Columns: {len(df_wide.columns)}")

    print("\nConsent Status:")
    consent_counts = df_wide['consent_status'].value_counts()
    for status, count in consent_counts.items():
        print(f"  - {status}: {count}")

    print(f"\nDosing Rescheduled: {df_wide['dosing_rescheduled'].sum()} participants")
    print(f"Participants with name: {df_wide['consent_nameprint'].notna().sum()}")

    print("\nData Availability by Scale (at any timepoint):")
    for scale in ['phq9_total', 'gad7_total', 'who5_total', 'psyflex_total',
                  'auditc_total', 'meq4_mean', 'ebi_total', 'piq_total', 'ceq_total']:
        cols = [c for c in df_wide.columns if c.startswith(f'{scale}_t')]
        if cols:
            any_data = df_wide[cols].notna().any(axis=1).sum()
            print(f"  - {scale}: {any_data} participants")

    print("\n" + "=" * 80)
    print("Integration complete!")
    print("=" * 80)

    return df_wide


if __name__ == '__main__':
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = Path(__file__).parent.parent / 'data' / 'sample_data.xlsx'

    if len(sys.argv) > 2:
        output_directory = Path(sys.argv[2])
    else:
        output_directory = Path(input_file).parent

    print(f"Input file: {input_file}")
    print(f"Output directory: {output_directory}\n")

    df_analytic = integrate_full(input_file, output_directory)

    print("\nPreview (first 15 participants):")
    cols = ['record_id', 'consent_nameprint', 'consent_status', 'dosing_rescheduled', 'age', 'n_events', 'timepoints']
    cols = [c for c in cols if c in df_analytic.columns]
    print(df_analytic[cols].head(15).to_string())
